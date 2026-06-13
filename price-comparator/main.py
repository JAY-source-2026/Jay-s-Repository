#!/usr/bin/env python3
"""단가 비교 프로그램 - 기준 단가와 거래명세서의 단가를 비교합니다."""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import re
import base64
import io
from pathlib import Path

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    from docx import Document
    DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_SUPPORT = True
except ImportError:
    DND_SUPPORT = False

try:
    from pdf2image import convert_from_path
    import pytesseract
    OCR_SUPPORT = True
except ImportError:
    OCR_SUPPORT = False

ITEM_KEYWORDS  = ['제품명', '품목', '품명', '상품명', '품번', '코드', '항목', '내역', 'item', 'product', 'name', 'code']
SPEC_KEYWORDS  = ['규격', 'size', 'spec']
PRICE_KEYWORDS = ['단가', '가격', '단위금액', '공급가', '판매가', 'price']
INVOICE_PAGE_KWS = ['거래명세서', '거래 명세서', '거래명세표']

COMPANY_COLOR = '#2d5682'



def _normalize_col_name(name):
    """PDF 컬럼명의 한글 자간 공백 제거: '단 가' → '단가', '규 격' → '규격'."""
    name = re.sub(r'(?<=[가-힯])\s+(?=[가-힯(])', '', str(name))
    return name.strip()


def detect_columns(df):
    item_col = spec_col = price_col = None
    for col in df.columns:
        col_str = _normalize_col_name(str(col)).lower()
        if item_col  is None and any(kw in col_str for kw in ITEM_KEYWORDS):
            item_col  = col
        if spec_col  is None and any(kw in col_str for kw in SPEC_KEYWORDS):
            spec_col  = col
        if price_col is None and any(kw in col_str for kw in PRICE_KEYWORDS):
            price_col = col
    return item_col, spec_col, price_col


def parse_excel(filepath):
    _HDR_KWS  = ['제품명', '품목', '품명', '단가', 'UNIT PRICE', 'PRICE', 'SIZE', 'AMOUNT', "Q'TY", 'REMARK']
    _STOP_KWS = ['TOTAL', '합계', '특기사항', '운송비', '익월말', '자재산출']

    dfs = {}
    xl = pd.ExcelFile(filepath)
    for sheet in xl.sheet_names:
        # ── 1단계: header=None 으로 읽어 실제 헤더 행 탐색 ──────────
        raw = pd.read_excel(filepath, sheet_name=sheet, header=None, dtype=str)
        raw = raw.fillna('')

        best_row, best_score = 0, 0
        for i in range(min(40, len(raw))):
            text = ' '.join(str(v) for v in raw.iloc[i])
            score = sum(1 for kw in _HDR_KWS if kw in text)
            if score > best_score:
                best_score, best_row = score, i

        # ── 2단계: 감지된 헤더 행으로 재읽기 ────────────────────────
        header_row = best_row if best_score >= 2 else 0
        df = pd.read_excel(filepath, sheet_name=sheet, header=header_row, dtype=str)

        # Unnamed 컬럼(병합셀 잔재) 이름 정리
        df.columns = [
            c if not str(c).startswith('Unnamed:') else ''
            for c in df.columns
        ]

        df = df.dropna(how='all').reset_index(drop=True)

        # ── 3단계: TOTAL/합계 이후 행 제거 ──────────────────────────
        stop_idx = None
        for i, row in df.iterrows():
            text = ' '.join(str(v) for v in row if str(v) not in ('nan', ''))
            if any(kw.upper() in text.upper() for kw in _STOP_KWS):
                stop_idx = i
                break
        if stop_idx is not None:
            df = df.iloc[:stop_idx]

        df = df.dropna(how='all').reset_index(drop=True)
        if not df.empty:
            dfs[sheet] = df
    return dfs


def _words_to_df(words):
    """x/y 좌표 기반으로 단어를 표로 재구성."""
    from collections import defaultdict

    HEADER_KWS = ['제품명', '품목', '품명', '상품명', '내역', '규격', '단위', '수량', '단가', '금액',
                  'SIZE', 'UNIT', "Q'TY", 'AMOUNT', 'REMARK', 'PRICE']
    STOP_KWS   = ['TOTAL', '합계', '특기사항', '운송비', '익월말', '현금', '자재산출', '담당자']

    # y좌표 기준으로 행 그룹핑 (15px 단위 — 미세하게 어긋난 셀도 같은 행으로 묶음)
    row_map = defaultdict(list)
    for w in words:
        y = round(w['top'] / 15) * 15
        row_map[y].append(w)

    sorted_ys = sorted(row_map.keys())

    # 헤더 행 탐색: HEADER_KWS 매칭 개수가 가장 많은 행 선택 (최소 3개 이상)
    best_y, best_score = None, 0
    for y in sorted_ys:
        texts = ' '.join(w['text'] for w in row_map[y])
        score = sum(1 for kw in HEADER_KWS if kw in texts)
        if score > best_score:
            best_score, best_y = score, y

    if best_y is None or best_score < 2:
        return None

    header_y = best_y
    header_words = sorted(row_map[header_y], key=lambda w: w['x0'])

    # 전후 ±30px 이내 인접 헤더 행도 병합 (한글 헤더 + 영문 헤더가 살짝 다른 y에 있는 경우)
    for y in sorted_ys:
        if y == header_y:
            continue
        if abs(y - header_y) > 30:
            if y > header_y + 30:
                break
            continue
        next_texts = ' '.join(w['text'] for w in row_map[y])
        if sum(1 for kw in HEADER_KWS if kw in next_texts) >= 1:
            for nw in row_map[y]:
                cx = (nw['x0'] + nw['x1']) / 2
                if not any(abs((hw['x0'] + hw['x1']) / 2 - cx) < 25 for hw in header_words):
                    header_words.append(nw)
            header_words = sorted(header_words, key=lambda w: w['x0'])

    # 인접 단어 중 x 간격이 12px 이내인 것은 같은 셀 → 하나로 합치기
    # (PRODUCT NAME, UNIT PRICE, 단가, 규격 등 두 단어로 쪼개진 헤더 처리)
    merged_names   = []
    merged_centers = []
    j = 0
    while j < len(header_words):
        w    = header_words[j]
        name = _normalize_col_name(w['text'])
        x0   = w['x0']
        x1   = w['x1']
        while j + 1 < len(header_words):
            nw  = header_words[j + 1]
            if nw['x0'] - x1 < 12:
                name = _normalize_col_name(name + ' ' + nw['text'])
                x1   = nw['x1']
                j   += 1
            else:
                break
        merged_names.append(name)
        merged_centers.append((x0 + x1) / 2)
        j += 1

    col_names   = merged_names
    col_centers = merged_centers

    boundaries = [0]
    for i in range(len(col_centers) - 1):
        boundaries.append((col_centers[i] + col_centers[i + 1]) / 2)
    boundaries.append(9999)

    def assign_col(x):
        for i in range(len(boundaries) - 1):
            if boundaries[i] <= x < boundaries[i + 1]:
                return i
        return len(boundaries) - 2

    # 데이터 시작 y 결정 (헤더 이후 첫 행)
    data_start_y = next((y for y in sorted_ys if y > header_y + 8), None)
    if data_start_y is None:
        return None

    # 데이터 행 구성
    rows_data = []
    pending = [''] * len(col_names)   # 직전 행과 병합 대기 중인 행

    def flush(r):
        if any(c.strip() for c in r):
            rows_data.append(r[:])

    for y in sorted_ys:
        if y < data_start_y:
            continue

        line_words = sorted(row_map[y], key=lambda w: w['x0'])
        line_text  = ' '.join(w['text'] for w in line_words)

        if any(kw in line_text.upper() for kw in STOP_KWS):
            break

        row = [''] * len(col_names)
        for w in line_words:
            cx = (w['x0'] + w['x1']) / 2
            ci = assign_col(cx)
            if ci < len(row):
                row[ci] = (row[ci] + ' ' + w['text']).strip()

        if all(c == '' for c in row):
            continue

        # 첫 컬럼(제품명)이 비어 있고 직전 행 첫 컬럼에 값이 있으면 같은 행으로 병합
        if row[0] == '' and pending[0] != '':
            for ci in range(1, len(row)):
                if row[ci]:
                    pending[ci] = (pending[ci] + ' ' + row[ci]).strip()
        else:
            flush(pending)
            pending = row

    flush(pending)

    if not rows_data:
        return None

    df = pd.DataFrame(rows_data, columns=col_names, dtype=str)
    df = df[df.apply(lambda r: any(v.strip() for v in r), axis=1)].reset_index(drop=True)
    return df if not df.empty else None


def _clean_table_to_df(table, key):
    """pdfplumber 테이블 데이터를 DataFrame으로 변환."""
    if not table or len(table) < 2:
        return None
    # 헤더 정리
    raw_header = [str(h).strip() if h else '' for h in table[0]]
    # 헤더가 전부 비어있으면 첫 행을 헤더로 쓸 수 없음 → 컬럼명 자동 생성
    if all(h == '' for h in raw_header):
        raw_header = [f'Col{k}' for k in range(len(raw_header))]
    # 중복 헤더 처리
    seen = {}
    header = []
    for h in raw_header:
        h = h or f'Col{len(header)}'
        if h in seen:
            seen[h] += 1
            header.append(f'{h}_{seen[h]}')
        else:
            seen[h] = 0
            header.append(h)
    df = pd.DataFrame(table[1:], columns=header, dtype=str)
    df = df.dropna(how='all').reset_index(drop=True)
    # 데이터가 있는 행만 유지 (전체 셀이 None/빈칸인 행 제거)
    df = df[df.apply(lambda r: any(str(v).strip() not in ('', 'None') for v in r), axis=1)]
    return df if not df.empty else None


def _merge_ocr_words(words, y_tol=30, x_gap_max=45):
    """OCR 개별 자모를 인접 좌표 기준으로 단어로 병합, 테두리 문자 제거."""
    from collections import defaultdict
    NOISE = re.compile(r'^[-—|=+\\]+$')

    row_map = defaultdict(list)
    for w in words:
        if NOISE.match(w['text']):
            continue
        y_key = round(w['top'] / y_tol) * y_tol
        row_map[y_key].append(w)

    merged = []
    for y_key in sorted(row_map.keys()):
        row = sorted(row_map[y_key], key=lambda w: w['x0'])
        cur = None
        for w in row:
            if cur is None:
                cur = dict(w)
            elif w['x0'] - cur['x1'] <= x_gap_max:
                cur['text'] = cur['text'] + w['text']
                cur['x1'] = max(cur['x1'], w['x1'])
                cur['bottom'] = max(cur['bottom'], w['bottom'])
            else:
                if cur['text'].strip():
                    merged.append(cur)
                cur = dict(w)
        if cur and cur['text'].strip():
            merged.append(cur)
    return merged


def _ocr_words_from_image(img):
    """pytesseract로 이미지에서 단어와 좌표를 추출 (pdfplumber words 포맷 호환)."""
    data = pytesseract.image_to_data(img, lang='kor+eng', output_type=pytesseract.Output.DICT)
    words = []
    for idx in range(len(data['text'])):
        text = data['text'][idx].strip()
        if not text or int(data['conf'][idx]) < 20:
            continue
        words.append({
            'text': text,
            'x0':    float(data['left'][idx]),
            'x1':    float(data['left'][idx] + data['width'][idx]),
            'top':   float(data['top'][idx]),
            'bottom': float(data['top'][idx] + data['height'][idx]),
        })
    return _merge_ocr_words(words)


def _parse_ocr_invoice_page(words):
    """거래명세서 OCR 단어에서 품목·단가 데이터를 추출하여 DataFrame 반환.

    각 행이 날짜(MM/DD)로 시작하는 패턴을 이용.
    오른쪽 숫자 3개 = 단가·공급가액·부가세로 매핑.
    """
    from collections import defaultdict
    DATE_PAT = re.compile(r'^\d{1,2}/\d{1,2}$')

    row_map = defaultdict(list)
    for w in words:
        y = round(w['top'] / 40) * 40   # 40px: OCR y오차 흡수용
        row_map[y].append(w)

    records = []
    for y in sorted(row_map.keys()):
        row_words = sorted(row_map[y], key=lambda w: w['x0'])
        tokens = [w['text'] for w in row_words]

        if not tokens or not DATE_PAT.match(tokens[0]):
            continue

        prices, labels = [], []
        for t in tokens[1:]:
            clean = re.sub(r'[^\d]', '', t)
            if clean and len(clean) >= 3 and not re.search(r'[가-힯a-zA-Z]', t):
                prices.append(clean)
            elif t.strip():
                labels.append(t)

        if not labels:
            continue

        item = labels[0]
        spec = labels[1] if len(labels) > 1 else ''
        # 마지막 3개 숫자 = 단가, 공급가액, 부가세
        unit_price = prices[-3] if len(prices) >= 3 else (prices[0] if prices else '')
        supply     = prices[-2] if len(prices) >= 2 else ''
        vat        = prices[-1] if prices else ''

        records.append({'품목': item, '규격': spec, '단가': unit_price,
                        '공급가액': supply, '부가세': vat})

    return pd.DataFrame(records) if records else None


def _ocr_words_from_page(filepath, page_index):
    """스캔 PDF 페이지를 이미지로 변환 후 OCR, pdfplumber words 포맷으로 반환."""
    if not OCR_SUPPORT:
        return []
    try:
        images = convert_from_path(
            filepath, dpi=300,
            first_page=page_index + 1,
            last_page=page_index + 1,
        )
        if not images:
            return []
        return _ocr_words_from_image(images[0])
    except Exception:
        return []


def parse_pdf(filepath):
    if not PDF_SUPPORT:
        return {}

    # pdfplumber 표 추출 전략 목록 (다양한 PDF 구조 대응)
    strategies = [
        {"vertical_strategy": "lines",  "horizontal_strategy": "lines"},
        {"vertical_strategy": "lines",  "horizontal_strategy": "lines",
         "snap_tolerance": 5, "join_tolerance": 5, "edge_min_length": 3},
        {"vertical_strategy": "lines",  "horizontal_strategy": "lines",
         "snap_tolerance": 10, "join_tolerance": 10},
        {"vertical_strategy": "text",   "horizontal_strategy": "lines"},
        {"vertical_strategy": "lines",  "horizontal_strategy": "text"},
        {"vertical_strategy": "text",   "horizontal_strategy": "text",
         "snap_tolerance": 5},
        {},  # pdfplumber 기본값
    ]

    dfs = {}
    with pdfplumber.open(filepath) as pdf:
        for i, page in enumerate(pdf.pages):
            page_key = f'페이지{i+1}'
            found = False

            # ── 1순위: 좌표 기반 파싱 (견적서·명세서 등 비정형 표에 강함) ──
            try:
                words = page.extract_words(x_tolerance=3, y_tolerance=3) or []

                # 스캔 PDF: 텍스트 레이어가 거의 없으면 OCR 폴백
                ocr_used = False
                if len(words) < 10 and OCR_SUPPORT:
                    words = _ocr_words_from_page(filepath, i)
                    ocr_used = True

                if words:
                    page_text = ' '.join(w['text'] for w in words)
                    is_invoice = any(kw in page_text for kw in INVOICE_PAGE_KWS)

                    if ocr_used and is_invoice:
                        # 스캔 거래명세서: 날짜 행 기반 전용 파서
                        df = _parse_ocr_invoice_page(words)
                    else:
                        df = _words_to_df(words)

                    if df is not None:
                        key = f'{page_key}_거래명세서' if is_invoice else f'{page_key}_표1'
                        dfs[key] = df
                        found = True
            except Exception:
                pass

            # ── 2순위: pdfplumber 전략별 표 추출 ─────────────────────────
            if not found:
                for strategy in strategies:
                    try:
                        tables = page.extract_tables(strategy)
                        if not tables:
                            continue
                        for j, table in enumerate(tables):
                            df = _clean_table_to_df(table, f'{page_key}_표{j+1}')
                            if df is not None:
                                dfs[f'{page_key}_표{j+1}'] = df
                                found = True
                        if found:
                            break
                    except Exception:
                        continue

            # ── 3순위: 텍스트 공백 분할 파싱 ────────────────────────────
            if not found:
                try:
                    text = page.extract_text()
                    if not text:
                        continue
                    lines = [l.strip() for l in text.splitlines() if l.strip()]
                    if len(lines) < 2:
                        continue
                    rows = [re.split(r'\s{2,}|\t', line) for line in lines]
                    max_cols = max(len(r) for r in rows)
                    if max_cols < 2:
                        continue
                    header_row = next((r for r in rows if len(r) == max_cols), rows[0])
                    header = [c.strip() or f'Col{k}' for k, c in enumerate(header_row)]
                    data_rows = [r for r in rows if r != header_row]
                    padded = [r + [''] * (len(header) - len(r)) for r in data_rows]
                    df = pd.DataFrame(padded, columns=header, dtype=str)
                    df = df.dropna(how='all').reset_index(drop=True)
                    if not df.empty:
                        dfs[f'{page_key}_텍스트'] = df
                except Exception:
                    continue

    return dfs


def parse_word(filepath):
    if not DOCX_SUPPORT:
        return {}
    dfs = {}
    doc = Document(filepath)
    for i, table in enumerate(doc.tables):
        data = [[cell.text.strip() for cell in row.cells] for row in table.rows]
        if len(data) < 2:
            continue
        header = [str(h) if h else f'Col{k}' for k, h in enumerate(data[0])]
        df = pd.DataFrame(data[1:], columns=header, dtype=str)
        df = df.dropna(how='all').reset_index(drop=True)
        if not df.empty:
            dfs[f'표{i+1}'] = df
    return dfs


def parse_file(filepath):
    ext = Path(filepath).suffix.lower()
    if ext in ('.xlsx', '.xls'):
        return parse_excel(filepath)
    elif ext == '.pdf':
        return parse_pdf(filepath)
    elif ext in ('.docx', '.doc'):
        return parse_word(filepath)
    elif ext == '.csv':
        df = pd.read_csv(filepath, encoding='utf-8-sig', dtype=str)
        return {'Sheet1': df}
    return {}


def normalize_price(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).replace(',', '').replace('₩', '').replace('원', '').strip()
    try:
        return float(s)
    except Exception:
        return None


def normalize_item(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    return str(value).strip().lower().replace(' ', '').replace('\n', '')


class DropZone(tk.Label):
    """드래그앤드롭 가능한 파일 업로드 영역."""

    def __init__(self, parent, text, on_drop_callback, **kwargs):
        super().__init__(parent, text=text, font=('맑은 고딕', 10),
                         bg='#e8f4fd', relief='groove', bd=2,
                         cursor='hand2', wraplength=250, **kwargs)
        self._callback = on_drop_callback
        if DND_SUPPORT:
            self.drop_target_register(DND_FILES)
            self.dnd_bind('<<Drop>>', self._on_dnd)

    def _on_dnd(self, event):
        raw = event.data.strip()
        files = re.findall(r'\{([^}]+)\}|(\S+)', raw)
        paths = [f[0] or f[1] for f in files]
        self._callback(paths)


class PriceComparatorApp:
    def __init__(self):
        if DND_SUPPORT:
            self.root = TkinterDnD.Tk()
        else:
            self.root = tk.Tk()

        self.root.title("단가 비교 프로그램")
        self.root.geometry("1280x820")
        self.root.minsize(900, 650)
        self.root.configure(bg='#f5f6fa')

        self.ref_data = {}
        self.ref_file = None
        self.invoice_files = []  # list of (filepath, {sheet: df})
        self.results_data = []
        self._sort_reverse = {}   # col → bool

        self._build_ui()

    # ── UI 구성 ──────────────────────────────────────────────────

    def _build_ui(self):
        # 헤더
        hdr = tk.Frame(self.root, bg=COMPANY_COLOR, height=64)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)

        # 로고 이미지 로드
        try:
            from PIL import Image, ImageTk
            logo_path = Path(__file__).parent / 'doenc_logo.png'
            if logo_path.exists():
                img = Image.open(logo_path).resize((40, 40), Image.LANCZOS)
                # 흰색 배경 블렌드 (로고 PNG를 회사색 배경 위에 자연스럽게)
                bg = Image.new('RGBA', img.size, (*bytes.fromhex(COMPANY_COLOR[1:]), 255))
                merged = Image.alpha_composite(bg, img.convert('RGBA'))
                self._logo_img = ImageTk.PhotoImage(merged.convert('RGB'))
                tk.Label(hdr, image=self._logo_img, bg=COMPANY_COLOR).pack(
                    side='left', padx=(14, 4), pady=12)
        except Exception:
            pass

        tk.Label(hdr, text="단가 비교 프로그램", font=('맑은 고딕', 15, 'bold'),
                 bg=COMPANY_COLOR, fg='white').pack(side='left', padx=(4, 16), pady=12)
        tk.Label(hdr, text="made by DOENC J.M  ", font=('Consolas', 9),
                 bg=COMPANY_COLOR, fg='#a8c8e8').pack(side='right', pady=16)

        body = tk.Frame(self.root, bg='#f5f6fa')
        body.pack(fill='both', expand=True, padx=18, pady=12)

        # 상단: 두 패널
        top = tk.Frame(body, bg='#f5f6fa')
        top.pack(fill='x')
        top.columnconfigure(0, weight=1)
        top.columnconfigure(1, weight=1)

        self._build_ref_panel(top)
        self._build_inv_panel(top)

        # 버튼 행
        self._build_controls(body)

        # 결과 테이블
        self._build_results(body)

    def _card(self, parent, title, col):
        f = tk.LabelFrame(parent, text=f'  {title}  ', font=('맑은 고딕', 11, 'bold'),
                          bg='white', bd=1, relief='solid', padx=12, pady=10)
        f.grid(row=0, column=col, sticky='nsew', padx=(0 if col else 0, 8 if col == 0 else 0),
               pady=0)
        return f

    def _build_ref_panel(self, parent):
        card = self._card(parent, '기준 단가 문서', 0)
        card.grid(padx=(0, 6))

        self.ref_drop = DropZone(card,
            '파일을 여기에 드래그하거나\n클릭해서 열기\n\n(Excel / PDF / Word)',
            lambda paths: self._load_ref(paths[0]) if paths else None,
            height=7)
        self.ref_drop.pack(fill='x', pady=(0, 8))
        self.ref_drop.bind('<Button-1>', lambda e: self._open_ref_dialog())

        self.ref_status = tk.Label(card, text='파일 없음', font=('맑은 고딕', 9),
                                   bg='white', fg='#999')
        self.ref_status.pack(anchor='w')

        self.ref_item_var,  self.ref_item_cb  = self._col_selector(card, '품목 컬럼')
        self.ref_spec_var,  self.ref_spec_cb  = self._col_selector(card, '규격 컬럼')
        self.ref_price_var, self.ref_price_cb = self._col_selector(card, '단가 컬럼')

    def _build_inv_panel(self, parent):
        card = self._card(parent, '거래명세서 (여러 파일 가능)', 1)
        card.grid(padx=(6, 0))

        self.inv_drop = DropZone(card,
            '파일을 여기에 드래그하거나\n클릭해서 열기\n\n(여러 파일 동시 선택 가능)',
            self._add_invoice_files,
            height=7)
        self.inv_drop.pack(fill='x', pady=(0, 8))
        self.inv_drop.bind('<Button-1>', lambda e: self._open_inv_dialog())

        # 파일 리스트
        list_wrap = tk.Frame(card, bg='white')
        list_wrap.pack(fill='x', pady=(0, 6))

        self.inv_list = tk.Listbox(list_wrap, font=('맑은 고딕', 9), height=4,
                                   selectmode='single', bg='#fafafa', relief='flat', bd=1)
        sb = ttk.Scrollbar(list_wrap, orient='vertical', command=self.inv_list.yview)
        self.inv_list.configure(yscrollcommand=sb.set)
        self.inv_list.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        btn_row = tk.Frame(card, bg='white')
        btn_row.pack(fill='x', pady=(0, 8))
        self._btn(btn_row, '선택 제거', self._remove_inv_file, '#e17055').pack(side='left', padx=(0, 4))
        self._btn(btn_row, '전체 제거', self._clear_inv_files, '#d63031').pack(side='left')

        self.inv_item_var,  self.inv_item_cb  = self._col_selector(card, '품목 컬럼')
        self.inv_spec_var,  self.inv_spec_cb  = self._col_selector(card, '규격 컬럼')
        self.inv_price_var, self.inv_price_cb = self._col_selector(card, '단가 컬럼')

    def _col_selector(self, parent, label):
        row = tk.Frame(parent, bg='white')
        row.pack(fill='x', pady=2)
        tk.Label(row, text=f'{label}:', font=('맑은 고딕', 9), bg='white', width=8,
                 anchor='w').pack(side='left')
        var = tk.StringVar(value='(자동감지)')
        cb = ttk.Combobox(row, textvariable=var, state='disabled', width=20)
        cb.pack(side='left', padx=(4, 0))
        return var, cb

    def _btn(self, parent, text, cmd, color='#0984e3'):
        return tk.Button(parent, text=text, command=cmd, font=('맑은 고딕', 9),
                         bg=color, fg='white', relief='flat', padx=10, pady=4,
                         cursor='hand2', activebackground=color)

    def _build_controls(self, parent):
        bar = tk.Frame(parent, bg='#f5f6fa')
        bar.pack(fill='x', pady=10)

        self._btn(bar, '▶  비교 시작', self._run_comparison, '#00b894').pack(side='left', padx=(0, 10), ipady=4)
        self.export_btn = self._btn(bar, '결과 내보내기 (Excel)', self._export, '#0984e3')
        self.export_btn.pack(side='left', ipady=4)
        self.export_btn.config(state='disabled')

        self.status_lbl = tk.Label(bar, text='', font=('맑은 고딕', 10),
                                   bg='#f5f6fa', fg='#636e72')
        self.status_lbl.pack(side='left', padx=16)

    def _build_results(self, parent):
        card = tk.LabelFrame(parent, text='  비교 결과  ', font=('맑은 고딕', 11, 'bold'),
                             bg='white', bd=1, relief='solid', padx=10, pady=8)
        card.pack(fill='both', expand=True)

        # 요약
        # ── 요약 라벨 (테이블 행 색상과 일치) ──────────────────────
        summ = tk.Frame(card, bg='white')
        summ.pack(fill='x', pady=(0, 8))

        def _pill(parent, text, bg, fg):
            f = tk.Frame(parent, bg=bg, padx=10, pady=3)
            f.pack(side='left', padx=(0, 8))
            lbl = tk.Label(f, text=text, font=('맑은 고딕', 10, 'bold'), bg=bg, fg=fg)
            lbl.pack()
            return lbl

        self.lbl_total    = _pill(summ, '전체: 0건',     '#f0f0f0', '#2d3436')
        self.lbl_match    = _pill(summ, '✓ 일치: 0건',   '#d0f5e8', '#1a6b50')
        self.lbl_mismatch = _pill(summ, '✗ 불일치: 0건', '#ffd6d6', '#8b0000')
        self.lbl_notfound = _pill(summ, '? 단가없음: 0건','#fff0e0', '#8b4513')

        # ── 트리뷰 ────────────────────────────────────────────────
        self._cols    = ('status', 'file', 'item', 'spec', 'ref_price', 'inv_price')
        self._headers = ('상태', '파일명', '품목명', '규격', '기준 단가', '거래 단가')
        widths        = (90, 180, 200, 180, 110, 110)

        wrap = tk.Frame(card)
        wrap.pack(fill='both', expand=True)

        self.tree = ttk.Treeview(wrap, columns=self._cols, show='headings')
        for col, hd, w in zip(self._cols, self._headers, widths):
            self.tree.heading(col, text=hd + '  ↕',
                              command=lambda c=col: self._sort_col(c))
            self.tree.column(col, width=w, anchor='center', minwidth=50)

        ysb = ttk.Scrollbar(wrap, orient='vertical',   command=self.tree.yview)
        xsb = ttk.Scrollbar(wrap, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        ysb.grid(row=0, column=1, sticky='ns')
        xsb.grid(row=1, column=0, sticky='ew')
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)

        self.tree.tag_configure('match',    background='#d0f5e8')
        self.tree.tag_configure('mismatch', background='#ffd6d6')
        self.tree.tag_configure('notfound', background='#fff0e0')

    # ── 파일 로딩 ─────────────────────────────────────────────────

    def _open_ref_dialog(self):
        fp = filedialog.askopenfilename(
            title='기준 단가 문서 선택',
            filetypes=[('지원 파일', '*.xlsx *.xls *.pdf *.docx *.csv'), ('모든 파일', '*.*')])
        if fp:
            self._load_ref(fp)

    def _load_ref(self, filepath):
        try:
            data = parse_file(filepath)
            if not data:
                messagebox.showerror('오류', '파일을 읽을 수 없습니다.\n지원 형식: Excel, PDF, Word, CSV')
                return
            self.ref_data = data
            self.ref_file = filepath
            first_df = next(iter(data.values()))
            cols = list(first_df.columns)
            item_col, spec_col, price_col = detect_columns(first_df)
            cols_with_none = ['(없음)'] + cols

            self.ref_drop.config(text=f'✓  {Path(filepath).name}', bg='#d0f5e8')
            self.ref_status.config(text=f'{len(data)}개 시트  /  {sum(len(d) for d in data.values())}행', fg='#00b894')

            self.ref_item_cb['values']  = cols
            self.ref_spec_cb['values']  = cols_with_none
            self.ref_price_cb['values'] = cols
            self.ref_item_cb.config(state='readonly')
            self.ref_spec_cb.config(state='readonly')
            self.ref_price_cb.config(state='readonly')
            self.ref_item_var.set(item_col or cols[0])
            self.ref_spec_var.set(spec_col or '(없음)')
            self.ref_price_var.set(price_col or cols[-1])
        except Exception as e:
            messagebox.showerror('오류', f'파일 읽기 실패:\n{e}')

    def _open_inv_dialog(self):
        fps = filedialog.askopenfilenames(
            title='거래명세서 선택 (여러 파일 가능)',
            filetypes=[('지원 파일', '*.xlsx *.xls *.pdf *.docx *.csv'), ('모든 파일', '*.*')])
        self._add_invoice_files(list(fps))

    def _add_invoice_files(self, paths):
        for fp in paths:
            if any(p == fp for p, _ in self.invoice_files):
                continue
            try:
                data = parse_file(fp)
                if not data:
                    messagebox.showerror('오류', f'읽을 수 없음:\n{Path(fp).name}')
                    continue
                self.invoice_files.append((fp, data))
                self.inv_list.insert('end', f'  ✓  {Path(fp).name}')

                # 첫 파일 기준으로 컬럼 자동 감지
                if len(self.invoice_files) == 1:
                    first_df = next(iter(data.values()))
                    cols = list(first_df.columns)
                    item_col, spec_col, price_col = detect_columns(first_df)
                    cols_with_none = ['(없음)'] + cols
                    self.inv_item_cb['values']  = cols
                    self.inv_spec_cb['values']  = cols_with_none
                    self.inv_price_cb['values'] = cols
                    self.inv_item_cb.config(state='readonly')
                    self.inv_spec_cb.config(state='readonly')
                    self.inv_price_cb.config(state='readonly')
                    self.inv_item_var.set(item_col or cols[0])
                    self.inv_spec_var.set(spec_col or '(없음)')
                    self.inv_price_var.set(price_col or cols[-1])
            except Exception as e:
                messagebox.showerror('오류', f'파일 읽기 실패:\n{e}')

    def _remove_inv_file(self):
        sel = self.inv_list.curselection()
        if sel:
            idx = sel[0]
            self.inv_list.delete(idx)
            self.invoice_files.pop(idx)

    def _clear_inv_files(self):
        self.inv_list.delete(0, 'end')
        self.invoice_files.clear()

    # ── 정렬 ──────────────────────────────────────────────────────

    def _sort_col(self, col):
        reverse = self._sort_reverse.get(col, False)
        col_idx = self._cols.index(col)

        def sort_key(iid):
            val = self.tree.item(iid, 'values')[col_idx]
            try:
                return (0, float(str(val).replace(',', '')))
            except ValueError:
                return (1, str(val))

        items = sorted(self.tree.get_children(''), key=sort_key, reverse=reverse)
        for i, iid in enumerate(items):
            self.tree.move(iid, '', i)

        self._sort_reverse[col] = not reverse
        arrow = '↓' if reverse else '↑'
        for c, hd in zip(self._cols, self._headers):
            mark = f'  {arrow}' if c == col else '  ↕'
            self.tree.heading(c, text=hd + mark)

    # ── 비교 로직 ─────────────────────────────────────────────────

    def _run_comparison(self):
        if not self.ref_data:
            messagebox.showwarning('경고', '기준 단가 문서를 먼저 선택해주세요.')
            return
        if not self.invoice_files:
            messagebox.showwarning('경고', '거래명세서를 선택해주세요.')
            return

        r_item  = self.ref_item_var.get()
        r_spec  = self.ref_spec_var.get()
        r_price = self.ref_price_var.get()
        i_item  = self.inv_item_var.get()
        i_spec  = self.inv_spec_var.get()
        i_price = self.inv_price_var.get()

        use_spec = (r_spec not in ('(없음)', '') and i_spec not in ('(없음)', ''))

        def make_key(item_val, spec_val=''):
            item = normalize_item(item_val) or ''
            spec = normalize_item(spec_val) or '' if use_spec else ''
            return f'{item}||{spec}' if spec else item

        # 기준 단가 딕셔너리 구축
        ref_prices = {}
        for df in self.ref_data.values():
            if r_item not in df.columns or r_price not in df.columns:
                continue
            for _, row in df.iterrows():
                spec_val  = row[r_spec]  if use_spec and r_spec  in df.columns else ''
                key       = make_key(row[r_item], spec_val)
                price     = normalize_price(row[r_price])
                if key and price is not None:
                    ref_prices[key] = price

        if not ref_prices:
            messagebox.showerror('오류', '기준 단가 문서에서 데이터를 찾을 수 없습니다.\n컬럼 설정을 확인해주세요.')
            return

        for row in self.tree.get_children():
            self.tree.delete(row)
        self.results_data = []

        total = match = mismatch = notfound = 0

        for filepath, data in self.invoice_files:
            fname = Path(filepath).name
            for df in data.values():
                if i_item not in df.columns or i_price not in df.columns:
                    continue
                for _, row in df.iterrows():
                    spec_val  = row[i_spec]  if use_spec and i_spec  in df.columns else ''
                    key       = make_key(row[i_item], spec_val)
                    inv_price = normalize_price(row[i_price])
                    if not key or inv_price is None:
                        continue

                    total += 1
                    ref_price = ref_prices.get(key)

                    orig_name = str(row[i_item]).strip()
                    orig_spec = str(row[i_spec]).strip() if use_spec and i_spec in df.columns else ''

                    if ref_price is None:
                        status = '? 단가 없음'
                        tag    = 'notfound'
                        notfound += 1
                        vals = (status, fname, orig_name, orig_spec, '단가 없음', f'{inv_price:,.0f}')
                    else:
                        if abs(inv_price - ref_price) < 1:
                            status, tag = '✓ 일치', 'match'
                            match += 1
                        else:
                            status, tag = '✗ 불일치', 'mismatch'
                            mismatch += 1
                        vals = (status, fname, orig_name, orig_spec,
                                f'{ref_price:,.0f}', f'{inv_price:,.0f}')

                    self.tree.insert('', 'end', values=vals, tags=(tag,))
                    self.results_data.append(vals)

        self.lbl_total.config(text=f'전체: {total}건')
        self.lbl_match.config(text=f'✓ 일치: {match}건')
        self.lbl_mismatch.config(text=f'✗ 불일치: {mismatch}건')
        self.lbl_notfound.config(text=f'? 단가없음: {notfound}건')

        if total == 0:
            messagebox.showwarning('결과 없음', '비교할 데이터가 없습니다.\n컬럼 설정을 확인해주세요.')
        else:
            self.export_btn.config(state='normal')
            self.status_lbl.config(
                text=f'비교 완료  |  {mismatch}건 불일치  /  {total}건 전체',
                fg='#d63031' if mismatch else '#00b894')

    # ── 내보내기 ──────────────────────────────────────────────────

    def _export(self):
        if not self.results_data:
            return
        fp = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel 파일', '*.xlsx')],
            title='결과 저장')
        if not fp:
            return

        headers = ['상태', '파일명', '품목명', '규격', '기준 단가', '거래 단가']
        df = pd.DataFrame(self.results_data, columns=headers)

        with pd.ExcelWriter(fp, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='비교결과')
            ws = writer.sheets['비교결과']

            from openpyxl.styles import PatternFill, Font, Alignment
            fills = {
                '✓ 일치':   PatternFill('solid', fgColor='D0F5E8'),
                '✗ 불일치': PatternFill('solid', fgColor='FFD6D6'),
                '? 미발견': PatternFill('solid', fgColor='FFF0E0'),
            }
            bold = Font(bold=True)

            # 헤더 스타일
            for cell in ws[1]:
                cell.font = bold
                cell.alignment = Alignment(horizontal='center')

            # 데이터 색상
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                status = row[0].value or ''
                fill = next((v for k, v in fills.items() if k in status), None)
                if fill:
                    for cell in row:
                        cell.fill = fill

            # 열 너비 자동 조정
            for col in ws.columns:
                w = max(len(str(c.value or '')) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(w + 4, 45)

        messagebox.showinfo('저장 완료', f'결과가 저장되었습니다:\n{fp}')

    def run(self):
        self.root.mainloop()


if __name__ == '__main__':
    app = PriceComparatorApp()
    app.run()
