"""단가 비교 프로그램 - Streamlit 웹 버전"""

import io
import re
import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill

from main import (
    detect_columns,
    normalize_item,
    normalize_price,
    parse_file,
)

COMPANY_COLOR = "#2d5682"

st.set_page_config(
    page_title="단가 비교 프로그램",
    page_icon="📊",
    layout="wide",
)

# ── 스타일 ──────────────────────────────────────────────────────────────────
st.markdown(
    f"""
    <style>
    /* 헤더 */
    .app-header {{
        background: {COMPANY_COLOR};
        padding: 16px 24px;
        border-radius: 8px;
        margin-bottom: 20px;
        display: flex;
        align-items: center;
        gap: 12px;
    }}
    .app-header h1 {{
        color: white;
        font-size: 22px;
        margin: 0;
        font-weight: 700;
    }}
    .app-header span {{
        color: #a8c8e8;
        font-size: 12px;
        margin-left: auto;
    }}

    /* 결과 테이블 */
    .match    {{ background-color: #d0f5e8 !important; }}
    .mismatch {{ background-color: #ffd6d6 !important; }}
    .notfound {{ background-color: #fff0e0 !important; }}

    /* 요약 뱃지 */
    .badge {{
        display: inline-block;
        padding: 4px 14px;
        border-radius: 20px;
        font-weight: 700;
        font-size: 14px;
        margin-right: 8px;
    }}
    .badge-total    {{ background:#f0f0f0; color:#2d3436; }}
    .badge-match    {{ background:#d0f5e8; color:#1a6b50; }}
    .badge-mismatch {{ background:#ffd6d6; color:#8b0000; }}
    .badge-notfound {{ background:#fff0e0; color:#8b4513; }}

    /* 섹션 카드 */
    section[data-testid="stSidebar"] {{ display: none; }}
    </style>
    """,
    unsafe_allow_html=True,
)

# ── 헤더 ────────────────────────────────────────────────────────────────────
logo_path = Path(__file__).parent / "doenc_logo.png"
col_logo, col_title = st.columns([1, 11])
with col_logo:
    if logo_path.exists():
        st.image(str(logo_path), width=48)
with col_title:
    st.markdown(
        f'<div style="background:{COMPANY_COLOR};padding:12px 20px;border-radius:8px;">'
        f'<span style="color:white;font-size:20px;font-weight:700;">단가 비교 프로그램</span>'
        f'<span style="color:#a8c8e8;font-size:11px;float:right;line-height:28px;">made by DOENC J.M</span>'
        f"</div>",
        unsafe_allow_html=True,
    )

st.markdown("---")


# ── 유틸 ─────────────────────────────────────────────────────────────────────
def save_uploaded(uploaded) -> str:
    """업로드 파일을 임시 파일로 저장 후 경로 반환."""
    suffix = Path(uploaded.name).suffix
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(uploaded.read())
    tmp.flush()
    return tmp.name


def build_excel_bytes(df: pd.DataFrame) -> bytes:
    """결과 DataFrame을 색상 적용 Excel 파일로 변환."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="비교결과")
        ws = writer.sheets["비교결과"]

        fills = {
            "✓ 일치":   PatternFill("solid", fgColor="D0F5E8"),
            "✗ 불일치": PatternFill("solid", fgColor="FFD6D6"),
            "? 단가 없음": PatternFill("solid", fgColor="FFF0E0"),
        }
        bold = Font(bold=True)

        for cell in ws[1]:
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status = row[0].value or ""
            fill = next((v for k, v in fills.items() if k in status), None)
            if fill:
                for cell in row:
                    cell.fill = fill

        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 45)

    return buf.getvalue()


# ── 파일 업로드 패널 ──────────────────────────────────────────────────────────
col_ref, col_inv = st.columns(2)

with col_ref:
    st.subheader("📄 기준 단가 문서")
    ref_file = st.file_uploader(
        "Excel / PDF / Word / CSV",
        type=["xlsx", "xls", "pdf", "docx", "csv"],
        key="ref",
    )

with col_inv:
    st.subheader("🧾 거래명세서")
    inv_files = st.file_uploader(
        "여러 파일 동시 업로드 가능",
        type=["xlsx", "xls", "pdf", "docx", "csv"],
        accept_multiple_files=True,
        key="inv",
    )

# ── 파일 파싱 ─────────────────────────────────────────────────────────────────
ref_data: dict = {}
inv_data: list = []  # list of (filename, {sheet: df})

if ref_file:
    with st.spinner("기준 단가 문서 읽는 중..."):
        try:
            path = save_uploaded(ref_file)
            ref_data = parse_file(path)
        except Exception as e:
            st.error(f"기준 단가 문서 읽기 실패: {e}")

if inv_files:
    with st.spinner("거래명세서 읽는 중..."):
        for uf in inv_files:
            try:
                path = save_uploaded(uf)
                data = parse_file(path)
                if data:
                    inv_data.append((uf.name, data))
                else:
                    st.warning(f"읽기 실패: {uf.name}")
            except Exception as e:
                st.warning(f"{uf.name} 오류: {e}")

# ── 컬럼 선택 ─────────────────────────────────────────────────────────────────
if ref_data or inv_data:
    st.markdown("---")
    st.subheader("⚙️ 컬럼 설정")
    col_r, col_i = st.columns(2)

    def col_selector(label, data, key_prefix):
        if not data:
            return None, None, None
        first_df = next(iter(data.values()))
        cols = list(first_df.columns)
        item_col, spec_col, price_col = detect_columns(first_df)
        cols_with_none = ["(없음)"] + cols

        c1, c2, c3 = st.columns(3)
        with c1:
            item = st.selectbox(f"{label} 품목", cols,
                                index=cols.index(item_col) if item_col in cols else 0,
                                key=f"{key_prefix}_item")
        with c2:
            spec = st.selectbox(f"{label} 규격", cols_with_none,
                                index=cols_with_none.index(spec_col) if spec_col in cols_with_none else 0,
                                key=f"{key_prefix}_spec")
        with c3:
            price = st.selectbox(f"{label} 단가", cols,
                                 index=cols.index(price_col) if price_col in cols else len(cols) - 1,
                                 key=f"{key_prefix}_price")
        return item, spec, price

    with col_r:
        if ref_data:
            st.markdown(f"**기준 단가** ({ref_file.name if ref_file else ''})")
            r_item, r_spec, r_price = col_selector("기준", ref_data, "ref")
        else:
            r_item = r_spec = r_price = None

    with col_i:
        if inv_data:
            st.markdown(f"**거래명세서** ({len(inv_data)}개 파일)")
            first_inv_data = inv_data[0][1]
            i_item, i_spec, i_price = col_selector("명세서", first_inv_data, "inv")
        else:
            i_item = i_spec = i_price = None

# ── 비교 실행 ─────────────────────────────────────────────────────────────────
st.markdown("---")

run = st.button("▶  비교 시작", type="primary", disabled=not (ref_data and inv_data))

if run and ref_data and inv_data:
    use_spec = r_spec not in ("(없음)", "", None) and i_spec not in ("(없음)", "", None)

    def make_key(item_val, spec_val=""):
        item = normalize_item(item_val) or ""
        spec = normalize_item(spec_val) or "" if use_spec else ""
        return f"{item}||{spec}" if spec else item

    # 기준 단가 딕셔너리
    ref_prices: dict = {}
    for df in ref_data.values():
        if r_item not in df.columns or r_price not in df.columns:
            continue
        for _, row in df.iterrows():
            spec_val = row[r_spec] if use_spec and r_spec in df.columns else ""
            key = make_key(row[r_item], spec_val)
            price = normalize_price(row[r_price])
            if key and price is not None:
                ref_prices[key] = price

    if not ref_prices:
        st.error("기준 단가 문서에서 데이터를 찾을 수 없습니다. 컬럼 설정을 확인해주세요.")
        st.stop()

    rows = []
    for fname, data in inv_data:
        for df in data.values():
            if i_item not in df.columns or i_price not in df.columns:
                continue
            for _, row in df.iterrows():
                spec_val = row[i_spec] if use_spec and i_spec in df.columns else ""
                key = make_key(row[i_item], spec_val)
                inv_price = normalize_price(row[i_price])
                if not key or inv_price is None:
                    continue

                orig_name = str(row[i_item]).strip()
                orig_spec = str(row[i_spec]).strip() if use_spec and i_spec in df.columns else ""
                ref_price = ref_prices.get(key)

                if ref_price is None:
                    status = "? 단가 없음"
                    ref_disp = "단가 없음"
                elif abs(inv_price - ref_price) < 1:
                    status = "✓ 일치"
                    ref_disp = f"{ref_price:,.0f}"
                else:
                    status = "✗ 불일치"
                    ref_disp = f"{ref_price:,.0f}"

                rows.append({
                    "상태": status,
                    "파일명": fname,
                    "품목명": orig_name,
                    "규격": orig_spec,
                    "기준 단가": ref_disp,
                    "거래 단가": f"{inv_price:,.0f}",
                })

    if not rows:
        st.warning("비교할 데이터가 없습니다. 컬럼 설정을 확인해주세요.")
        st.stop()

    result_df = pd.DataFrame(rows)
    total     = len(result_df)
    match     = (result_df["상태"] == "✓ 일치").sum()
    mismatch  = (result_df["상태"] == "✗ 불일치").sum()
    notfound  = (result_df["상태"] == "? 단가 없음").sum()

    # 요약 뱃지
    st.markdown(
        f'<span class="badge badge-total">전체: {total}건</span>'
        f'<span class="badge badge-match">✓ 일치: {match}건</span>'
        f'<span class="badge badge-mismatch">✗ 불일치: {mismatch}건</span>'
        f'<span class="badge badge-notfound">? 단가없음: {notfound}건</span>',
        unsafe_allow_html=True,
    )
    st.markdown("")

    # 색상 적용 테이블
    def row_color(row):
        s = row["상태"]
        if s == "✓ 일치":
            return ["background-color: #d0f5e8"] * len(row)
        if s == "✗ 불일치":
            return ["background-color: #ffd6d6"] * len(row)
        if s == "? 단가 없음":
            return ["background-color: #fff0e0"] * len(row)
        return [""] * len(row)

    styled = result_df.style.apply(row_color, axis=1)
    st.dataframe(styled, use_container_width=True, hide_index=True)

    # Excel 내보내기
    excel_bytes = build_excel_bytes(result_df)
    st.download_button(
        label="📥  결과 Excel 다운로드",
        data=excel_bytes,
        file_name="단가비교결과.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
